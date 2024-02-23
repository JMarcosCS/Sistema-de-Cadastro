import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl
import xlrd
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side


# Setando a aparencia padrão
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.todo_sistema()

    def layout_config(self):
        self.title("Gestão Deepsea")
        self.geometry("700x500")

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(
            self, text="Tema", bg_color="transparent", text_color=["#000", "#fff",]).place(x=50, y=430)
        self.opt_apm = ctk.CTkOptionMenu(
            self, values=["Light", "Dark", "System"], command=self.change_apm).place(x=50, y=460)

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0,
                             bg_color="teal", fg_color="teal").place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Gestão Deepsea", font=(
            "Century Gohtic bold", 24), bg_color="teal", text_color="#000").place(x=250, y=22)
        span = ctk.CTkLabel(frame, text="Por favor, preencher todo o formulário!", font=(
            "Century Gohtic bold", 16), text_color=["#000", "#FFF"]).place(x=50, y=70)

        ficheiro = pathlib.Path("Clientes.xlsx")

        if ficheiro.exists():
            pass
        else:
            ficheiro = Workbook()
            folha = ficheiro.active
            folha["A1"] = "Nome Completo"
            folha["B1"] = "Contato"
            folha["C1"] = "Idade"
            folha["D1"] = "Gênero"
            folha["E1"] = "Endereço"
            folha["F1"] = "Observações"

            wb = openpyxl.Workbook()
            sheet = wb.active
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            my_style = Style(border=thin_border)    
            sheet.cell(row=1, column=1).border = my_style
            sheet.cell(row = 1, column = 1).font = Font(size = 12, bold=True)
            sheet.column_dimensions['A','B','C','D','E','F'].width = 20

            ficheiro.save("Clientes.xlsx")

        def submit():

            name = name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            gender = gender_combobox.get()
            address = address_value.get()
            obs = obs_entry.get(0.0, END)

            if (name == "" or contact == "" or age == "" or address == ""):
                messagebox.showerror(
                    "Sistema", "ERRO\nPor favor preencha todos os campos!")
            else:

                ficheiro = openpyxl.load_workbook("Clientes.xlsx")
                folha = ficheiro.active
                folha.cell(column=1, row=folha.max_row+1, value=name)
                folha.cell(column=2, row=folha.max_row, value=contact)
                folha.cell(column=3, row=folha.max_row, value=age)
                folha.cell(column=4, row=folha.max_row, value=gender)
                folha.cell(column=5, row=folha.max_row, value=address)
                folha.cell(column=6, row=folha.max_row, value=obs)

                ficheiro.save(r"CLientes.xlsx")
                messagebox.showinfo("Sistema", "Dados salvo com sucesso")

        def clear():
            name = name_value.set("")
            contact = contact_value.set("")
            age = age_value.set("")
            address = address_value.set("")
            obs = obs_entry.delete(0.0, END)

        # Text variables
        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        address_value = StringVar()

        # Entrys
        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=(
            "Century Gohtic", 16), fg_color="transparent")
        contact_entry = ctk.CTkEntry(self, width=200, textvariable=contact_value, font=(
            "Century Gohtic", 16), fg_color="transparent")
        age_entry = ctk.CTkEntry(self, width=150, textvariable=age_value, font=(
            "Century Gohtic", 16), fg_color="transparent")
        address_entry = ctk.CTkEntry(self, width=200, textvariable=address_value, font=(
            "Century Gohtic", 16), fg_color="transparent")

        # Combobox
        gender_combobox = ctk.CTkComboBox(
            self, values=["Masculino", "Feminino"], font=("Century Gohtic bold", 14))
        gender_combobox.set("Masculino")

        # Entrada de observações
        obs_entry = ctk.CTkTextbox(self, width=500, height=150, font=(
            "Arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")

        # Labels
        lb_name = ctk.CTkLabel(frame, text="Nome Completo", font=(
            "Century Gohtic bold", 16), text_color=["#000", "#FFF"])
        lb_contact = ctk.CTkLabel(frame, text="Contato", font=(
            "Century Gohtic bold", 16), text_color=["#000", "#FFF"])
        lb_age = ctk.CTkLabel(frame, text="Idade", font=(
            "Century Gohtic bold", 16), text_color=["#000", "#FFF"])
        lb_gender = ctk.CTkLabel(frame, text="Gênero", font=(
            "Century Gohtic bold", 16), text_color=["#000", "#FFF"])
        lb_address = ctk.CTkLabel(frame, text="Endereço", font=(
            "Century Gohtic bold", 16), text_color=["#000", "#FFF"])
        lb_obs = ctk.CTkLabel(frame, text="Observação", font=(
            "Century Gohtic bold", 16), text_color=["#000", "#FFF"])

        btn_submit = ctk.CTkButton(
            self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=300, y=420)
        btn_submit = ctk.CTkButton(
            self, text="Limpar campos".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=500, y=420)

        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_contact.place(x=450, y=120)
        contact_entry.place(x=450, y=150)

        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y=220)

        lb_gender.place(x=500, y=190)
        gender_combobox.place(x=500, y=220)

        lb_address.place(x=50, y=190)
        address_entry.place(x=50, y=220)

        lb_obs.place(x=50, y=260)
        obs_entry.place(x=150, y=260)

    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)


if __name__ == "__main__":
    app = App()
    app.mainloop()
